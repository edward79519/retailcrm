from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from retailCRM.settings import BASE_DIR


def readxlsx(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    try:
        rawdata = []
        for row_data in ws.iter_rows(min_col=1, max_col=7, min_row=2):
            rowdata = {}
            for col in row_data:
                label = ws["{}1".format(col.column_letter)].value
                rowdata[label] = col.value
            if rowdata['項次'] != '範例':
                rawdata.append(rowdata)
        if len(rawdata) != 0:
            data = {
                'status': 'OK',
                'rawdata': rawdata,
            }
        else:
            data = {
                'status': 'Error',
                'msg': "There is no data."
            }
        return data
    except:
        data = {
            'status': 'Error',
            'msg': 'Something wrong. Please contact administrator.'
        }
        return data